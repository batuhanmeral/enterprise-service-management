import csv
import io
import os
import sys
from collections import defaultdict
from datetime import date, datetime, time, timedelta

from django.contrib.auth.decorators import login_required
from django.db.models import (
    Avg, Count, DurationField, ExpressionWrapper, F, Q,
)
from django.db.models.functions import TruncDay, TruncWeek, TruncMonth
from django.http import HttpResponse, HttpResponseForbidden
from django.template.loader import render_to_string
from django.utils import timezone
from django.views.generic import TemplateView

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill

if sys.platform == 'win32':
    _gtk_bin = os.environ.get('GTK_RUNTIME_DIR', r'C:\Program Files\GTK3-Runtime Win64\bin')
    if os.path.isdir(_gtk_bin):
        os.add_dll_directory(_gtk_bin)

from weasyprint import HTML

from departments.models import Department
from identity.models import Role, User
from identity.views import ManagerOrAdminRequiredMixin
from tickets.models import Status, Ticket

from .filters import ReportFilterForm


# Sadece MANAGER veya ADMIN raporlara erişebilir
def _require_manager_or_admin(user):
    return user.is_authenticated and user.role in (Role.MANAGER, Role.ADMIN)


_TRUNC_MAP = {
    'day': TruncDay,
    'week': TruncWeek,
    'month': TruncMonth,
}

_LABEL_FORMAT = {
    'day': '%d.%m',
    'week': '%d.%m',
    'month': '%b %Y',
}


# Raporlama dashboard'u — sadece MANAGER ve ADMIN erişebilir
class ReportDashboardView(ManagerOrAdminRequiredMixin, TemplateView):
    template_name = 'reports/dashboard.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        user = self.request.user

        form = ReportFilterForm(self.request.GET or None, user=user)
        context['form'] = form
        context['is_admin'] = user.role == Role.ADMIN

        ticket_qs = form.apply(Ticket.objects.all(), user)

        context['base_query'] = form.as_ticket_query()

        status_counts = dict(
            ticket_qs.values_list('status').annotate(c=Count('id'))
        )
        total = sum(status_counts.values())
        context['total_tickets'] = total
        context['open_count'] = status_counts.get(Status.OPEN, 0)
        context['in_progress_count'] = status_counts.get(Status.IN_PROGRESS, 0)
        context['resolved_count'] = status_counts.get(Status.RESOLVED, 0)
        context['closed_count'] = status_counts.get(Status.CLOSED, 0)
        context['escalated_count'] = status_counts.get(Status.ESCALATED, 0)

        closed_qs = ticket_qs.filter(
            status=Status.CLOSED, closed_at__isnull=False
        )
        sla_total = 0
        sla_breach = 0
        dept_sla_total = defaultdict(int)
        dept_sla_breach = defaultdict(int)
        cat_sla_total = defaultdict(int)
        cat_sla_breach = defaultdict(int)
        agent_sla_total = defaultdict(int)
        agent_sla_breach = defaultdict(int)
        for row in closed_qs.values(
            'closed_at', 'sla_due_at', 'department_id', 'category_id', 'assigned_to_id'
        ):
            breached = bool(row['sla_due_at'] and row['closed_at'] > row['sla_due_at'])
            sla_total += 1
            if breached:
                sla_breach += 1
            did = row['department_id']
            dept_sla_total[did] += 1
            if breached:
                dept_sla_breach[did] += 1
            cid = row['category_id']
            if cid is not None:
                cat_sla_total[cid] += 1
                if breached:
                    cat_sla_breach[cid] += 1
            aid = row['assigned_to_id']
            if aid is not None:
                agent_sla_total[aid] += 1
                if breached:
                    agent_sla_breach[aid] += 1
        sla_met = sla_total - sla_breach
        context['sla_total'] = sla_total
        context['sla_met'] = sla_met
        context['sla_breach'] = sla_breach
        context['sla_compliance_pct'] = (
            round(sla_met / sla_total * 100, 1) if sla_total else None
        )

        reopened_count = ticket_qs.filter(reopen_count__gt=0).count()
        context['reopened_count'] = reopened_count
        context['reopen_rate_pct'] = (
            round(reopened_count / total * 100, 1) if total else None
        )

        csat_qs = ticket_qs.filter(status=Status.CLOSED, csat_rating__isnull=False)
        csat_count = csat_qs.count()
        csat_avg = csat_qs.aggregate(avg=Avg('csat_rating'))['avg']
        context['csat_average'] = round(csat_avg, 2) if csat_avg is not None else None
        context['csat_count'] = csat_count
        csat_dist = dict(
            csat_qs.values_list('csat_rating').annotate(c=Count('id'))
        )
        context['csat_histogram'] = [csat_dist.get(i, 0) for i in range(1, 6)]

        granularity = form.get_granularity()
        trend_labels, trend_opened, trend_closed = _build_trend(
            ticket_qs, granularity, form
        )
        context['trend_labels'] = trend_labels
        context['trend_opened'] = trend_opened
        context['trend_closed'] = trend_closed
        context['granularity'] = granularity

        context['anomalies'] = _detect_category_anomalies(ticket_qs)

        if user.role == Role.ADMIN:
            dept_qs = Department.objects.all()
        else:
            dept_qs = Department.objects.filter(pk=user.department_id)

        dept_status_rows = (
            ticket_qs.values('department_id')
            .annotate(
                total=Count('id'),
                open_c=Count('id', filter=Q(status=Status.OPEN)),
                in_progress_c=Count('id', filter=Q(status=Status.IN_PROGRESS)),
                resolved_c=Count('id', filter=Q(status=Status.RESOLVED)),
                closed_c=Count('id', filter=Q(status=Status.CLOSED)),
                escalated_c=Count('id', filter=Q(status=Status.ESCALATED)),
                avg_dur=Avg(
                    ExpressionWrapper(
                        F('closed_at') - F('created_at'),
                        output_field=DurationField(),
                    ),
                    filter=Q(status=Status.CLOSED, closed_at__isnull=False),
                ),
            )
        )
        dept_status_map = {r['department_id']: r for r in dept_status_rows}

        departments = []
        for dept in dept_qs.order_by('name'):
            stats = dept_status_map.get(dept.pk, {})
            total_d = stats.get('total', 0)
            avg_dur = stats.get('avg_dur')
            sla_t = dept_sla_total.get(dept.pk, 0)
            sla_b = dept_sla_breach.get(dept.pk, 0)
            departments.append({
                'pk': dept.pk,
                'name': dept.name,
                'total': total_d,
                'open': stats.get('open_c', 0),
                'in_progress': stats.get('in_progress_c', 0),
                'resolved': stats.get('resolved_c', 0),
                'closed': stats.get('closed_c', 0),
                'escalated': stats.get('escalated_c', 0),
                'avg_hours': round(avg_dur.total_seconds() / 3600, 1) if avg_dur else None,
                'sla_total': sla_t,
                'sla_breach': sla_b,
                'sla_compliance_pct': (
                    round((sla_t - sla_b) / sla_t * 100, 1) if sla_t else None
                ),
            })
        context['departments'] = departments

        context['dept_names'] = [d['name'] for d in departments]
        context['dept_open'] = [d['open'] for d in departments]
        context['dept_in_progress'] = [d['in_progress'] for d in departments]
        context['dept_closed'] = [d['closed'] for d in departments]
        context['dept_escalated'] = [d['escalated'] for d in departments]

        top_cat_rows = (
            ticket_qs.filter(category__isnull=False)
            .values('category_id', 'category__name', 'category__department__name')
            .annotate(ticket_count=Count('id'))
            .order_by('-ticket_count')[:10]
        )
        context['top_categories'] = [
            {
                'pk': r['category_id'],
                'name': r['category__name'],
                'department_name': r['category__department__name'] or '—',
                'ticket_count': r['ticket_count'],
            }
            for r in top_cat_rows
        ]

        for cat in context['top_categories']:
            sla_t = cat_sla_total.get(cat['pk'], 0)
            sla_b = cat_sla_breach.get(cat['pk'], 0)
            cat['sla_total'] = sla_t
            cat['sla_breach'] = sla_b
            cat['sla_compliance_pct'] = (
                round((sla_t - sla_b) / sla_t * 100, 1) if sla_t else None
            )

        if user.role == Role.ADMIN:
            user_qs = User.objects.filter(role=Role.AGENT, is_active=True)
        else:
            user_qs = User.objects.filter(
                role=Role.AGENT, is_active=True, department=user.department
            )

        agent_rows = (
            ticket_qs.filter(assigned_to__isnull=False)
            .values('assigned_to_id')
            .annotate(
                active=Count('id', filter=Q(status=Status.IN_PROGRESS)),
                resolved=Count('id', filter=Q(status=Status.RESOLVED)),
                closed=Count('id', filter=Q(status=Status.CLOSED)),
                escalated=Count('id', filter=Q(status=Status.ESCALATED)),
                reopened=Count('id', filter=Q(reopen_count__gt=0)),
                total=Count('id'),
                avg_dur=Avg(
                    ExpressionWrapper(
                        F('closed_at') - F('created_at'),
                        output_field=DurationField(),
                    ),
                    filter=Q(status=Status.CLOSED, closed_at__isnull=False),
                ),
                avg_csat=Avg('csat_rating', filter=Q(csat_rating__isnull=False)),
            )
        )
        agent_map = {r['assigned_to_id']: r for r in agent_rows}

        personnel = []
        for u in user_qs.select_related('department').order_by('first_name', 'last_name'):
            row = agent_map.get(u.pk, {})
            avg_dur = row.get('avg_dur')
            avg_csat = row.get('avg_csat')
            tot = row.get('total', 0)
            reopened = row.get('reopened', 0)
            sla_t = agent_sla_total.get(u.pk, 0)
            sla_b = agent_sla_breach.get(u.pk, 0)
            personnel.append({
                'pk': u.pk,
                'name': u.get_full_name() or u.username,
                'department': u.department.name if u.department else '—',
                'active': row.get('active', 0),
                'resolved': row.get('resolved', 0),
                'closed': row.get('closed', 0),
                'escalated': row.get('escalated', 0),
                'reopened': reopened,
                'reopen_rate_pct': round(reopened / tot * 100, 1) if tot else None,
                'avg_hours': round(avg_dur.total_seconds() / 3600, 1) if avg_dur else None,
                'avg_csat': round(avg_csat, 2) if avg_csat else None,
                'sla_compliance_pct': (
                    round((sla_t - sla_b) / sla_t * 100, 1) if sla_t else None
                ),
            })
        context['personnel'] = personnel

        csat_ranked = sorted(
            [p for p in personnel if p['avg_csat'] is not None],
            key=lambda p: -p['avg_csat'],
        )[:10]
        reopen_ranked = sorted(
            [p for p in personnel if p['reopen_rate_pct'] is not None and p['reopened'] > 0],
            key=lambda p: -p['reopen_rate_pct'],
        )[:10]
        context['personnel_csat_ranked'] = csat_ranked
        context['personnel_reopen_ranked'] = reopen_ranked

        scorecard_rows = (
            ticket_qs.filter(assigned_to__isnull=False, category__isnull=False)
            .values(
                'category_id',
                'category__name',
                'category__department__name',
                'assigned_to_id',
                'assigned_to__first_name',
                'assigned_to__last_name',
                'assigned_to__username',
            )
            .annotate(
                c=Count('id'),
                closed_c=Count('id', filter=Q(status=Status.CLOSED)),
                clean_c=Count('id', filter=Q(status=Status.CLOSED, reopen_count=0)),
                avg_dur=Avg(
                    ExpressionWrapper(
                        F('closed_at') - F('created_at'),
                        output_field=DurationField(),
                    ),
                    filter=Q(status=Status.CLOSED, closed_at__isnull=False),
                ),
            )
        )
        cat_cards = {}
        for r in scorecard_rows:
            cid = r['category_id']
            card = cat_cards.get(cid)
            if card is None:
                card = cat_cards[cid] = {
                    'category_id': cid,
                    'category_name': r['category__name'],
                    'department_name': r['category__department__name'] or '—',
                    'total': 0,
                    'agents': [],
                }
            card['total'] += r['c']
            closed_c = r['closed_c']
            card['agents'].append({
                'agent_id': r['assigned_to_id'],
                'agent_name': (
                    f"{r['assigned_to__first_name']} {r['assigned_to__last_name']}".strip()
                    or r['assigned_to__username']
                ),
                'count': r['c'],
                'avg_hours': (
                    round(r['avg_dur'].total_seconds() / 3600, 1) if r['avg_dur'] else None
                ),
                'correct_rate': (
                    round(r['clean_c'] / closed_c * 100, 1) if closed_c else None
                ),
            })
        for card in cat_cards.values():
            card['agents'].sort(key=lambda a: -a['count'])
        context['category_scorecards'] = sorted(
            cat_cards.values(), key=lambda c: -c['total']
        )[:10]

        outgoing_rows = (
            ticket_qs.filter(sender__department__isnull=False, category__isnull=False)
            .values(
                'sender__department__name',
                'category__name',
                'category__department__name',
            )
            .annotate(c=Count('id'))
            .order_by('-c')[:60]
        )
        context['outgoing_dept_matrix'] = [
            {
                'sender_dept': r['sender__department__name'],
                'category_name': r['category__name'],
                'category_dept': r['category__department__name'] or '—',
                'count': r['c'],
            }
            for r in outgoing_rows
        ]

        top_senders = (
            ticket_qs.values(
                'sender_id',
                'sender__first_name',
                'sender__last_name',
                'sender__username',
                'sender__department__name',
            )
            .annotate(c=Count('id'))
            .order_by('-c')[:10]
        )
        context['top_senders'] = [
            {
                'pk': r['sender_id'],
                'name': (
                    f"{r['sender__first_name']} {r['sender__last_name']}".strip()
                    or r['sender__username']
                ),
                'department': r['sender__department__name'] or '—',
                'count': r['c'],
            }
            for r in top_senders
        ]

        return context



def _parse_date(date_str):
    if not date_str:
        return None
    try:
        return date.fromisoformat(date_str)
    except ValueError:
        return None


def _build_trend(ticket_qs, granularity, form):
    trunc_cls = _TRUNC_MAP.get(granularity, TruncMonth)
    label_fmt = _LABEL_FORMAT.get(granularity, '%b %Y')
    now = timezone.now()

    data = form.safe_cleaned()
    date_from = data.get('date_from')
    date_to = data.get('date_to')

    if date_from and date_to:
        start_dt = timezone.make_aware(datetime.combine(date_from, time.min))
        end_dt = timezone.make_aware(datetime.combine(date_to, time.max))
    else:
        end_dt = now
        if granularity == 'day':
            start_dt = now - timedelta(days=14)
        elif granularity == 'week':
            start_dt = now - timedelta(weeks=8)
        else:
            start_dt = now - timedelta(days=180)

    opened_rows = (
        ticket_qs.filter(created_at__gte=start_dt, created_at__lte=end_dt)
        .annotate(b=trunc_cls('created_at'))
        .values('b').annotate(c=Count('id')).order_by('b')
    )
    closed_rows = (
        ticket_qs.filter(
            status=Status.CLOSED,
            closed_at__isnull=False,
            closed_at__gte=start_dt,
            closed_at__lte=end_dt,
        )
        .annotate(b=trunc_cls('closed_at'))
        .values('b').annotate(c=Count('id')).order_by('b')
    )

    opened_map = {r['b']: r['c'] for r in opened_rows}
    closed_map = {r['b']: r['c'] for r in closed_rows}

    buckets = _generate_buckets(start_dt, end_dt, granularity)
    labels = [b.strftime(label_fmt) for b in buckets]
    opened = [opened_map.get(b, 0) for b in buckets]
    closed = [closed_map.get(b, 0) for b in buckets]
    return labels, opened, closed


def _generate_buckets(start, end, granularity):
    start = timezone.localtime(start)
    end = timezone.localtime(end)
    buckets = []
    if granularity == 'day':
        cur = start.replace(hour=0, minute=0, second=0, microsecond=0)
        end_norm = end.replace(hour=0, minute=0, second=0, microsecond=0)
        while cur <= end_norm:
            buckets.append(cur)
            cur += timedelta(days=1)
    elif granularity == 'week':
        cur = start.replace(hour=0, minute=0, second=0, microsecond=0)
        cur -= timedelta(days=cur.weekday())
        end_norm = end.replace(hour=0, minute=0, second=0, microsecond=0)
        while cur <= end_norm:
            buckets.append(cur)
            cur += timedelta(weeks=1)
    else:
        cur = start.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
        end_norm = end.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
        while cur <= end_norm:
            buckets.append(cur)
            year = cur.year + (1 if cur.month == 12 else 0)
            month = 1 if cur.month == 12 else cur.month + 1
            cur = cur.replace(year=year, month=month)
    return buckets


def _detect_category_anomalies(ticket_qs, min_current=5, threshold_pct=40):
    now = timezone.now()
    cur_start = now - timedelta(days=7)
    prev_start = now - timedelta(days=14)

    cur = list(
        ticket_qs.filter(created_at__gte=cur_start, category__isnull=False)
        .values('category__name')
        .annotate(c=Count('id'))
    )
    prev = dict(
        ticket_qs.filter(
            created_at__gte=prev_start,
            created_at__lt=cur_start,
            category__isnull=False,
        )
        .values_list('category__name')
        .annotate(c=Count('id'))
    )

    alerts = []
    for row in cur:
        name = row['category__name']
        c = row['c']
        if c < min_current:
            continue
        p = prev.get(name, 0)
        if p == 0:
            alerts.append({
                'category': name, 'current': c, 'previous': 0,
                'change_pct': None, 'is_new_trend': True,
            })
        else:
            change = (c - p) / p * 100
            if change >= threshold_pct:
                alerts.append({
                    'category': name, 'current': c, 'previous': p,
                    'change_pct': round(change, 0), 'is_new_trend': False,
                })

    alerts.sort(key=lambda x: (-(x['change_pct'] or 9999), -x['current']))
    return alerts



def _get_ticket_export_data(user, request_get):
    form = ReportFilterForm(request_get or None, user=user)
    qs = form.apply(
        Ticket.objects.select_related(
            'sender', 'assigned_to', 'department', 'category',
        ).prefetch_related('tags').order_by('-created_at'),
        user,
    )

    rows = []
    for t in qs:
        rows.append({
            'id': t.pk,
            'subject': t.subject,
            'status': t.get_status_display(),
            'priority': t.get_priority_display(),
            'department': t.department.name if t.department else '—',
            'category': t.category.name if t.category else '—',
            'tags': ', '.join([tag.name for tag in t.tags.all()]) if t.tags.all() else '—',
            'sender': (t.sender.get_full_name() or t.sender.username) if t.sender else '—',
            'assigned_to': (t.assigned_to.get_full_name() or t.assigned_to.username) if t.assigned_to else '—',
            'created_at': t.created_at.strftime('%d.%m.%Y %H:%M'),
            'closed_at': t.closed_at.strftime('%d.%m.%Y %H:%M') if t.closed_at else '—',
            'resolution_note': t.resolution_note or '',
        })
    return rows


EXPORT_HEADERS = [
    'ID', 'Konu', 'Durum', 'Öncelik', 'Departman', 'Kategori', 'Etiketler',
    'Talep Sahibi', 'Üstlenen Personel', 'Oluşturulma', 'Kapatılma', 'Çözüm Notu',
]


@login_required
def export_csv(request):
    if not _require_manager_or_admin(request.user):
        return HttpResponseForbidden('Bu rapora erişim yetkiniz bulunmamaktadır.')
    rows = _get_ticket_export_data(request.user, request.GET)
    response = HttpResponse(content_type='text/csv; charset=utf-8')
    response['Content-Disposition'] = 'attachment; filename="bilet_raporu.csv"'
    response.write('﻿')
    writer = csv.writer(response)
    writer.writerow(EXPORT_HEADERS)
    for r in rows:
        writer.writerow([
            r['id'], r['subject'], r['status'], r['priority'],
            r['department'], r['category'], r['tags'], r['sender'], r['assigned_to'],
            r['created_at'], r['closed_at'], r['resolution_note'],
        ])
    return response


@login_required
def export_excel(request):
    if not _require_manager_or_admin(request.user):
        return HttpResponseForbidden('Bu rapora erişim yetkiniz bulunmamaktadır.')
    rows = _get_ticket_export_data(request.user, request.GET)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Bilet Raporu'

    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='212529', end_color='212529', fill_type='solid')

    for col_idx, header in enumerate(EXPORT_HEADERS, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')

    keys = [
        'id', 'subject', 'status', 'priority', 'department', 'category', 'tags',
        'sender', 'assigned_to', 'created_at', 'closed_at', 'resolution_note',
    ]
    for row_idx, r in enumerate(rows, 2):
        for col_idx, key in enumerate(keys, 1):
            ws.cell(row=row_idx, column=col_idx, value=r[key])

    col_widths = [6, 30, 10, 10, 18, 18, 20, 20, 20, 18, 18, 40]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(
        output.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename="bilet_raporu.xlsx"'
    return response


@login_required
def export_pdf(request):
    if not _require_manager_or_admin(request.user):
        return HttpResponseForbidden('Bu rapora erişim yetkiniz bulunmamaktadır.')
    rows = _get_ticket_export_data(request.user, request.GET)

    total = len(rows)
    open_count = sum(1 for r in rows if r['status'] == 'Açık')
    in_progress_count = sum(1 for r in rows if r['status'] == 'İşlemde')
    resolved_count = sum(1 for r in rows if r['status'] == 'Çözüldü')
    closed_count = sum(1 for r in rows if r['status'] == 'Kapandı')
    escalated_count = sum(1 for r in rows if r['status'] == 'Eskalasyon')

    html = render_to_string('reports/export_pdf.html', {
        'rows': rows,
        'total': total,
        'open_count': open_count,
        'in_progress_count': in_progress_count,
        'resolved_count': resolved_count,
        'closed_count': closed_count,
        'escalated_count': escalated_count,
        'generated_at': timezone.now().strftime('%d.%m.%Y %H:%M'),
    })

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="bilet_raporu.pdf"'
    HTML(string=html, base_url=request.build_absolute_uri('/')).write_pdf(response)
    return response
